#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
import pandas as pd
from openpyxl import load_workbook


class RACalculatorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Martensite - Retained Austenite Calculator")
        self.root.geometry("700x650")
        self.file_path = None
        self.setup_ui()

    def setup_ui(self):
        main = ttk.Frame(self.root, padding="10")
        main.pack(fill=tk.BOTH, expand=True)

        notebook = ttk.Notebook(main)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        notebook.add(self.create_calculator_tab(), text="Calculator")
        notebook.add(self.create_info_tab(), text="Information")

        status = ttk.Frame(main, relief=tk.SUNKEN, padding="2")
        status.pack(fill=tk.X, side=tk.BOTTOM, pady=(5, 0))
        self.status_label = ttk.Label(status, text="Ready", font=('Arial', 9))
        self.status_label.pack(side=tk.LEFT)

        ttk.Label(main, text="Developed by Mahmoud Elaraby", font=('Arial', 11)).pack()

    def create_calculator_tab(self):
        frame = ttk.Frame(self.root, padding="20")

        ttk.Label(frame, text="Calculate martensite fraction (Fm) and retained austenite (RA):",
                  font=('Arial', 11, 'italic')).pack(pady=(0, 20), anchor='w')

        # File Selection
        file_frame = ttk.LabelFrame(frame, text="File Selection", padding="15")
        file_frame.pack(fill=tk.X, pady=10)

        ttk.Label(file_frame, text="Selected file:", font=('Arial', 9, 'bold')).pack(anchor='w')
        self.file_label = ttk.Label(file_frame, text="No file selected", font=('Arial', 9), foreground='darkgray')
        self.file_label.pack(anchor='w', padx=(10, 0))

        ttk.Button(file_frame, text="Select Excel File", command=self.select_file, width=25).pack(pady=(5, 0))

        # Parameters
        params = ttk.LabelFrame(frame, text="Calculation Parameters", padding="15")
        params.pack(fill=tk.X, pady=10)

        tq_frame = ttk.Frame(params)
        tq_frame.pack(fill=tk.X, pady=5)
        ttk.Label(tq_frame, text="Quenching Temperature (°C):", font=('Arial', 10)).pack(side=tk.LEFT)
        self.tq_var = tk.StringVar(value="25")
        ttk.Entry(tq_frame, textvariable=self.tq_var, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Label(tq_frame, text="(Default: 25°C)", font=('Arial', 8, 'italic'), foreground='darkgray').pack(
            side=tk.LEFT)

        # Calculate
        calc = ttk.LabelFrame(frame, text="Run Calculation", padding="15")
        calc.pack(fill=tk.X, pady=10)

        ttk.Label(calc, text="Calculate Ms, Fm, and RA and append to Excel file.",
                  font=('Arial', 9), foreground='darkblue').pack(pady=(0, 10))

        self.calc_btn = ttk.Button(calc, text="Calculate and Save Results",
                                   command=self.calculate_and_save, width=30, state=tk.DISABLED)
        self.calc_btn.pack()

        return frame

    def create_info_tab(self):
        frame = ttk.Frame(self.root, padding="20")

        # Required Columns
        req = ttk.LabelFrame(frame, text="Required Excel Columns", padding="15")
        req.pack(fill=tk.X, pady=10)

        ttk.Label(req, text="Your Excel file must contain:", font=('Arial', 10, 'bold')).pack(anchor='w', pady=(0, 10))

        for col in ["Mass_fraction_C_in_FCC_A1", "Mass_fraction_N_in_FCC_A1", "Mass_fraction_Mn_in_FCC_A1",
                    "Mass_fraction_Si_in_FCC_A1", "Mass_fraction_Al_in_FCC_A1", "Mass_fraction_Cr_in_FCC_A1",
                    "FCC_A1_Fraction"]:
            ttk.Label(req, text=f"• {col}", font=('Arial', 9)).pack(anchor='w', padx=(20, 0))

        # Formulas
        formula = ttk.LabelFrame(frame, text="Calculation Formulas", padding="15")
        formula.pack(fill=tk.X, pady=10)

        formulas = [
            ("Ms (Martensite Start):", "Ms = 692 - 502·√(wC + 0.86·wN) - 37·wMn - 14·wSi + 20·wAl - 11·wCr"),
            ("Martensite Fraction:",
             "α = 0.0231 - 0.0105·wC, β = 1.4304 - 1.1836·wC + 0.7527·wC²\nFm = (1 - exp(-α·(Ms - Tq)^β)) · f_γ"),
            ("Retained Austenite:", "RA = f_γ - Fm")
        ]

        for title, text in formulas:
            ttk.Label(formula, text=title, font=('Arial', 9, 'bold')).pack(anchor='w')
            ttk.Label(formula, text=text, font=('Arial', 8)).pack(anchor='w', padx=(20, 0), pady=(2, 10))

        # References
        ref = ttk.LabelFrame(frame, text="References", padding="15")
        ref.pack(fill=tk.X, pady=10)
        ttk.Label(ref,
                  text="[1] https://doi.org/10.1002/srin.202100576\n[2] Processing, microstructure and mechanical behavior of medium manganese steels.",
                  font=('Arial', 8, 'italic'), foreground='darkblue').pack(anchor='w')

        return frame

    def select_file(self):
        path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.file_path = path
            self.file_label.config(text=path.split('/')[-1], foreground='black')
            self.calc_btn.config(state=tk.NORMAL)
            self.status_label.config(text=f"File loaded: {path.split('/')[-1]}")

    def calculate_and_save(self):
        if not self.file_path:
            return messagebox.showwarning("Warning", "Select an Excel file first.")

        try:
            Tq = float(self.tq_var.get())
        except ValueError:
            return messagebox.showerror("Invalid Input", "Enter a valid quenching temperature.")

        try:
            self.status_label.config(text="Reading Excel file...")
            self.root.update_idletasks()

            df = pd.read_excel(self.file_path)

            required = ["Mass_fraction_C_in_FCC_A1", "Mass_fraction_N_in_FCC_A1", "Mass_fraction_Mn_in_FCC_A1",
                        "Mass_fraction_Si_in_FCC_A1", "Mass_fraction_Al_in_FCC_A1", "Mass_fraction_Cr_in_FCC_A1",
                        "FCC_A1_Fraction"]

            missing = [c for c in required if c not in df.columns]
            if missing:
                self.status_label.config(text="Ready")
                return messagebox.showerror("Missing Columns", "Missing:\n" + "\n".join(f"• {c}" for c in missing))

            self.status_label.config(text="Calculating...")
            self.root.update_idletasks()

            wC, wN, wMn, wSi, wAl, wCr = df[required[:6]].values.T

            Ms = 692 - 502 * np.sqrt(wC + 0.86 * wN) - 37 * wMn - 14 * wSi + 20 * wAl - 11 * wCr
            alpha = 0.0231 - 0.0105 * wC
            beta = 1.4304 - 1.1836 * wC + 0.7527 * wC ** 2

            f_gamma = df["FCC_A1_Fraction"]
            Fm = np.where(Ms > Tq, (1 - np.exp(-alpha * (Ms - Tq) ** beta)) * f_gamma, 0)
            RA = f_gamma - Fm

            self.status_label.config(text="Saving...")
            self.root.update_idletasks()

            wb = load_workbook(self.file_path)
            ws = wb.active
            start = ws.max_column + 1

            for i, (col, vals) in enumerate(zip(["Ms", "Fm", "RA"], [Ms, Fm, RA])):
                ws.cell(1, start + i, col)
                for j, v in enumerate(vals, 2):
                    ws.cell(j, start + i, v)

            wb.save(self.file_path)

            self.status_label.config(text="Completed!")
            messagebox.showinfo("Success", f"Calculated {len(df)} rows\nTq: {Tq}°C\nColumns added: Ms, Fm, RA")
            self.root.after(3000, lambda: self.status_label.config(text="Ready"))

        except Exception as e:
            self.status_label.config(text="Error")
            messagebox.showerror("Error", f"Error: {str(e)}")
            self.root.after(3000, lambda: self.status_label.config(text="Ready"))


def main():
    root = tk.Tk()
    RACalculatorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()